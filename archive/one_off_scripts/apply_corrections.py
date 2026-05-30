
"""
Apply corrections from a corrections JSON file (or from the reviewed Excel).
Backs up repertoire.db first.
"""
import sqlite3
import shutil
import json
import sys
from datetime import datetime
from pathlib import Path

DB_PATH = 'repertoire.db'
BACKUP_PATH = f'repertoire.db.backup_corrections_{datetime.now().strftime("%Y%m%d_%H%M%S")}'

# Default: load corrections from JSON file
input_file = sys.argv[1] if len(sys.argv) > 1 else 'corrections.json'

if input_file.endswith('.xlsx'):
    from openpyxl import load_workbook
    wb = load_workbook(input_file)
    corrections = []
    # Read Updates sheet
    ws = wb['Updates']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        corrections.append({
            'id': row[0],
            'action': 'update',
            'new_title': row[5],
            'new_composer': row[7],
            'new_dates': row[9],
        })
    # Read Deletions sheet
    ws = wb['Deletions']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        corrections.append({'id': row[0], 'action': 'delete'})
else:
    with open(input_file) as f:
        corrections = json.load(f)

# Show summary
n_d = sum(1 for c in corrections if c['action'] == 'delete')
n_u = sum(1 for c in corrections if c['action'] == 'update')
print(f"\nWill apply: {n_u} updates, {n_d} deletions")

ans = input("Proceed? [y/N] ").strip().lower()
if ans != 'y':
    print("Aborted.")
    sys.exit(0)

# Backup
shutil.copy(DB_PATH, BACKUP_PATH)
print(f"Backed up {DB_PATH} → {BACKUP_PATH}")

conn = sqlite3.connect(DB_PATH)
applied_u = 0
applied_d = 0
for c in corrections:
    if c['action'] == 'delete':
        conn.execute("DELETE FROM pieces WHERE id = ?", (c['id'],))
        applied_d += 1
    elif c['action'] == 'update':
        conn.execute(
            "UPDATE pieces SET title=?, composer=?, composer_dates=? WHERE id=?",
            (c.get('new_title', ''), c.get('new_composer', ''), c.get('new_dates', ''), c['id'])
        )
        applied_u += 1
conn.commit()
conn.close()

print(f"\nApplied {applied_u} updates, {applied_d} deletions.")
print(f"To restore: cp {BACKUP_PATH} {DB_PATH}")
